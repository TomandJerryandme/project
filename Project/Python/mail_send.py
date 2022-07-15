import win32com.client as win32
import pandas as pd
from pandas import DataFrame
from openpyxl import load_workbook
from collections import OrderedDict
# pdb为断点程序
import pdb

def test():
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    # 收件人添加
    # mail.Recipients.Add('1728703711@qq.com')
    # mail.Recipients.Add('liuxuan02@inspur.com')
    # mail.Recipients.Add('yanhuaqiang@inspur.com')
    # 收件人
    # 该属性是个字符串，但是可以以分号 ; 区分不同的邮箱，与 CC 属性一致
    # mail.To = 'liuxuan02@inspur.com'
    # 抄送人
    # mail.CC = 'wanglijuan01@inspur.com; niyulu@inspur.com'
    # mail.CC = 'liuxuan02@inspur.com'
    # 选择发送邮箱,只需要修改对应使用的邮箱/账号地址即可
    # mail.SentOnBehalfOfName = 'accoumt2@outlook.com'
    # 密抄收件人
    mail.Bcc='1728703711@qq.com'
    # 邮件主题
    mail.Subject = '邮件主题内容'
    # 邮件正文
    message = """
    尊敬的客户，你好：
        我是Outlook的客服人员，收到该邮件不用回复!
        该邮件是为了测试多个抄送人是否发送成功
    """
    hide_message = """
    <h6 style="color: white;">隐藏信息</h6>
    """
    # 获取签名，必须在对邮件正文赋值之前进行获取，获取后，签名存放在 mail.HTMLBody 里面
    mail.GetInspector
    # 将签名与正文放在一起，签名要放在后面
    mail.HTMLBody = message + hide_message + mail.HTMLBody
    mail.BodyFormat = 2
    # 设置重要性为高
    mail.Importance = 2
    # 添加附件 与收件人一样，可以多次添加
    # mail.Attachments.Add(r'D:\back\Download\all_summary.xlsx')
    # mail.HtmlBody = "<div><img src='D:\back\test.jpg'></div>"
    # 调用Outlook显示
    mail.Display()
    # 发送
    mail.Send()

def send_mail(title, receiver_list, message_body, cc_list=None, attachment_list=None):
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    # 收件人添加
    for receiver in receiver_list: mail.Recipients.Add(receiver)
    # 收件人
    # mail.To = 'liuxuan02@inspur.com'
    # 抄送人
    cc_str = '; '.join(cc_list) if cc_list else ''
    mail.CC = cc_str
    # mail.CC = 'liuxuan02@inspur.com'
    # 选择发送邮箱,只需要修改对应使用的邮箱/账号地址即可
    # mail.SentOnBehalfOfName = 'accoumt2@outlook.com'
    # 密抄收件人
    # mail.Bcc = '1728703711@qq.com'
    # 邮件主题
    mail.Subject = title
    # 邮件正文
    message = message_body
    # inSuite 财务周报 .xlsx
    # inSuite产品部周报-平台开发组（2022年6月份).xlsx
    # 获取签名，必须在对邮件正文赋值之前进行获取，获取后，签名存放在 mail.HTMLBody 里面
    mail.GetInspector
    # 将签名与正文放在一起，签名要放在后面
    mail.HTMLBody = message + mail.HTMLBody
    mail.BodyFormat = 2
    # 设置重要性为高
    mail.Importance = 2
    # 添加附件 与收件人一样，可以多次添加
    # mail.Attachments.Add(r'D:\back\Download\all_summary.xlsx')
    # mail.HtmlBody = "<div><img src='D:\back\test.jpg'></div>"
    # 调用Outlook显示
    mail.Display()
    # 发送
    # mail.Send()

def get_platform_mail_body(file_path):
    """
        @desc: 通过读取excel文件，组织发送邮件的body
    """
    this_week_dict = {}
    next_week_dict = {}
    file = load_workbook(file_path)
    query_sheet = [sheet for sheet in file.worksheets if sheet.sheet_state != 'hidden'][0]
    flag = '0'
    max_row = query_sheet.max_row
    max_column = query_sheet.max_column
    person_column = 0
    task_column = 0
    task = ''
    row = 0
    number_seq = 1
    row_list = []
    person_dict = OrderedDict()
    for x in range(2, max_row):
        if query_sheet.cell(x, 1).value == '三、本周工作总结':
            flag = '1'
            row = x
            row_list.append(row+1)
            continue
        if query_sheet.cell(x, 1).value == '四、下周工作计划':
            flag = '2'
            row = x
            row_list.append(row+1)
            continue
        if query_sheet.cell(x, 1).value == '五、上周问题跟踪':
            break
        # flag置为了True，可以进行读数据了
        if flag == '1' and (not task_column or not person_column):
            for y in range(1, max_column):
                if query_sheet.cell(x, y).value == '任务':
                    task_column = y
                    continue
                if query_sheet.cell(x, y).value == '责任人':
                    person_column = y
                    break
        if task_column != 0 and person_column != 0 and x not in row_list:
            task = query_sheet.cell(x, task_column).value
            person = query_sheet.cell(x, person_column).value
            seq = str(query_sheet.cell(x, number_seq).value)
            if person:
                if not person_dict.get(person, None):
                    person_dict[person] = person
                if not this_week_dict.get(person):
                    this_week_dict[person] = ""
                if not next_week_dict.get(person):
                    next_week_dict[person] = ""
                if flag == '1':
                    this_week_dict[person] += " " + seq + " " + task + "<br>"
                elif flag == '2':
                    next_week_dict[person] += " " + seq + " " + task + "<br>"
    message = ""
    for person in person_dict:
        # 编辑表体数据
        message += person + ':<br>本周任务：<br>' + this_week_dict[person] + '<br>' + '下周任务：<br>' + next_week_dict[person] + '<br>'
    return message

def get_fin_mail_body(file_path):
    """
        @desc: 通过读取excel文件，组织发送邮件的body
    """
    this_week_dict = {}
    next_week_dict = {}
    file = load_workbook(file_path)
    query_sheet = [sheet for sheet in file.worksheets if sheet.sheet_state != 'hidden'][0]
    flag = '0'
    max_row = query_sheet.max_row
    max_column = query_sheet.max_column
    person_column = 0
    task_column = 0
    task = ''
    row = 0
    number_seq = 1
    row_list = []
    person_dict = OrderedDict()
    for x in range(2, max_row):
        if query_sheet.cell(x, 1).value == '三、本周工作总结':
            flag = '1'
            row = x
            row_list.append(row+1)
            continue
        if query_sheet.cell(x, 1).value == '四、下周工作计划':
            flag = '2'
            row = x
            row_list.append(row+1)
            continue
        if query_sheet.cell(x, 1).value == '五、已解决问题列表':
            break
        # flag置为了True，可以进行读数据了
        if flag == '1' and (not task_column or not person_column):
            for y in range(1, max_column):
                if query_sheet.cell(x, y).value == '任务':
                    task_column = y
                    continue
                if query_sheet.cell(x, y).value == '责任人':
                    person_column = y
                    break
        if task_column != 0 and person_column != 0 and x not in row_list:
            # 可以取数了
            task = query_sheet.cell(x, task_column).value
            person = query_sheet.cell(x, person_column).value
            seq = str(query_sheet.cell(x, number_seq).value)
            if person:
                if not person_dict.get(person, None):
                    person_dict[person] = person
                if not this_week_dict.get(person):
                    this_week_dict[person] = ""
                if not next_week_dict.get(person):
                    next_week_dict[person] = ""
                task = task.replace('\n', '<br>&emsp;')
                if flag == '1':
                    this_week_dict[person] += "&emsp;" + " " + task + "<br>"
                elif flag == '2':
                    next_week_dict[person] += "&emsp;" + " " + task + "<br>"
    message = "<center><h1>周报</h1></center>"
    for person in person_dict:
        # 编辑表体数据
        message += person + ':<br> 本周任务：<br>' + this_week_dict[person] + '<br>' + ' 下周任务：<br>' + next_week_dict[person] + '<br>'
    return message

def get_mail_info():
    # inSuite 财务周报 .xlsx
    # inSuite产品部周报-平台开发组（2022年6月份).xlsx
    file_path = r'D:\ChromeDownload\\' + 'inSuite 财务周报 .xlsx'
    pdb.set_trace()
    message = get_fin_mail_body(file_path)
    title = '财务组周会纪要'
    # fin_person_list = ['wubai@inspur.com', 'limtr@inspur.com', 'songmx@inspur.com', 'fanjh@inspur.com', 'wangjiwen01@inspur.com', 'zhouysh@inspur.com', 'zhaofy_lc@inspur.com', 'yanhuaqiang@inspur.com', 'niyulu@inspur.com', 'yangxingcai@inspur.com', 'zhaibingbing@inspur.com', 'wutielin@inspur.com', 'zhanganjie@inspur.com', 'wanglijuan01@inspur.com']
    platfrom_person_list = ['wubai@inspur.com', 'sun_peng_lc@inspur.com', 'anqi@inspur.com', 'zhangwei04@inspur.com', 'shibinbin@inspur.com', 'zhangwei20@inspur.com', 'zhumengmeng@inspur.com', 'xushipeng@inspur.com', 'zhanganjie@inspur.com', 'sunhui03@inspur.com', 'sunqibing@inspur.com', 'liuxin04@inspur.com']
    # cc_list = ['zhengwb@inspur.com', 'chenman-s@inspur.com']
    fin_person_list = ['liuxuan02@inspur.com']
    cc_list = []
    return title, fin_person_list, message, cc_list, None

title, receiver_list, message_body, cc_list, attachment_list = get_mail_info()
send_mail(title, receiver_list, message_body, cc_list, attachment_list)
