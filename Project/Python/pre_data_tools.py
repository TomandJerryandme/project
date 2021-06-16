# -*- coding: utf-8 -*-

import re
from openpyxl import load_workbook
from read_excel import read_excel_file_to_generate_dict


KEY_VALUE_DICT = {
    "is_project_ctrl": "Boolean",
    "reimb_item_option": "Selection",
    "exp_item_ids": "Many2many",
    "is_support_currency": "Boolean",
}
COMMON_STR = """    <!-- {document_type_name}：{params_setting_type} -->
    <record id="{record_id}" model="mdm.document.params">
      <field name="module_id" search="[('name', '=', 'ps_er')]"></field>
      <field name="corespond_doc_id" ref="{document_list_id}"/>
      <field name="document_type_id" ref="{document_type_record_id}"/>
      <field name="biz_type_id" ref="{biz_type_id}"/>
      {key_value_str}
    </record>

"""
KEY_VALUE_STR = """<field name="key">{key_str}</field>
      <field name="value">{value_str}</field>
      <field name="data_type">{data_type_str}</field>"""

ITEMS_DICT = read_excel_file_to_generate_dict()

def generate_xml_data():
    file = load_workbook(r"C:\Users\liuxuan02\Desktop\document.xlsx")
    sheet = file["Sheet1"]
    mapp = {
        "是": "True",
        "否": "False",
    }
    for x in range(2, 68):  # 2-68
        document_type_name = sheet.cell(row=x, column=1).value
        document_type_number = sheet.cell(row=x, column=6).value
        is_support_currency = mapp.get(sheet.cell(row=x, column=3).value)
        is_project_ctrl = mapp.get(sheet.cell(row=x, column=4).value)
        biz_type_id = sheet.cell(row=x, column=8).value
        document_list_id = ""
        if document_type_number < 120:
            document_type_record_id = "er_travel_mdm_document_type_nyjk_000" + str(document_type_number)
            document_list_id = "er_travel_mdm_document_list"
        else:
            document_type_record_id = "er_expense_mdm_document_type_nyjk_000" + str(document_type_number)
            document_list_id = "er_expense_mdm_document_list"
        expense_items_str = sheet.cell(row=x, column=5).value
        expense_items_value = ''
        if expense_items_str:
            expense_items = expense_items_str.split(";")
            expense_items_value = """'_'.join([{items}])"""
            items_str = ", ".join("str(ref('ps_mdm.{id}'))".format(id=ITEMS_DICT.get(item)) for item in expense_items)
            expense_items_value = expense_items_value.format(items=items_str)
        dddd = {
            "is_support_currency": [is_support_currency, "支持外币业务"],
            "is_project_ctrl": [is_project_ctrl, "项目管控"],
            "reimb_item_option": ["include", "可选报销事项"],
            "exp_item_ids": [None, "可选报销事项范围"]
        }
        VALUE_INDEX = 0
        TRAN_INDEX = 1
        record_str_list = ""
        for key, value in KEY_VALUE_DICT.items():
            if key != "exp_item_ids":
                key_value_str = KEY_VALUE_STR.format(key_str=key, value_str=dddd.get(key)[VALUE_INDEX], data_type_str=value)
            else:
                key_value_str = KEY_VALUE_STR.format(key_str=key, value_str=expense_items_value, data_type_str=value)
            record_str = COMMON_STR.format(document_type_name=document_type_name, params_setting_type=dddd.get(key)[TRAN_INDEX],
                                           record_id="er_empense_{number}_mdm_document_params_{params}".format(number=str(document_type_number), params=key),
                                           document_type_record_id=document_type_record_id,
                                           key_value_str=key_value_str,
                                           biz_type_id=biz_type_id,
                                           document_list_id=document_list_id)
            record_str_list = record_str_list + record_str
        print(record_str_list)


generate_xml_data()