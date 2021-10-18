# 这个文件里主要测试或者记录一些对excel文件进行操作的方法

import  pandas  as pd
from pandas import DataFrame

#写
dic1 = {'标题列1': ['张三','李四'],
        '标题列2': [80, 90]
       }
df = pd.DataFrame(dic1)
df.to_excel('1.xlsx', index=False)



import json

import xlrd, xlwt
from openpyxl import load_workbook
#  E
def read_excel_file():
    file = load_workbook(r"C:\Users\liuxuan02\Desktop\data.xlsx")
    sheet = file["AAAA"]
    for x in range(2, 33):
        model = sheet.cell(row=x, column=1).value
        re = model.replace(".", "_")
        print('select * from %s;' % re)
        # sheet.cell(row=x, column=2).value = re

    file.save("data.xlsx")


#
# file = load_workbook(r"C:\Users\liuxuan02\Desktop\document_type.xlsx")
# sheet = file["Sheet1"]
# value = sheet.cell(row=3, column=12).value
# print(value)

def read_excel_file_to_generate_dict():
    file = load_workbook(r"C:\Users\liuxuan02\Desktop\dict_items.xlsx")
    sheet = file["Sheet1"]
    source_list = ""
    dict = {}
    for x in range(1, 91):
        key_value = sheet.cell(row=x, column=3).value
        value_value = sheet.cell(row=x, column=2).value
        dict[key_value] = value_value
    return dict

# read_excel_file_to_generate_dict()

# read_excel_file()


# read_excel_file()
# import pandas as pd
# file = open(r"C:\Users\liuxuan02\Desktop\document_type.xlsx", "rb")
# pd.read_excel(r"C:\Users\liuxuan02\Desktop\document_type.xlsx", )
