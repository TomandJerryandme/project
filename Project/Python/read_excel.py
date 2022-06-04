# -*- coding: utf-8 -*-

import json
import os.path
import xlrd, xlwt
from openpyxl import load_workbook
import xlsxwriter
import pandas as pd
from pandas import DataFrame

from time_counter import time_counter
#  E
def read_excel_file(file_path, fold_name, file_name, select_sheet_name):
    generate_dir = file_path + '\\' + fold_name
    if not os.path.exists(generate_dir):
        # 生成文件不存在，创建
        os.mkdir(generate_dir)
    load_file = file_path + "\\" + file_name
    file = load_workbook(load_file)
    sheet = file[select_sheet_name]
    split_number = 500
    tran_list = ""
    source_list = ""
    data_row = 1
    max_row = sheet.max_row
    max_column = sheet.max_column
    return
    for x in range(2, sheet.max_row):
        # 41列数据，10735条数据
        file_number = (x - 1) // 500
        file_name = generate_dir + '\大指库存初始导入模板%s-%s.xlsx' % (file_number * 500 + 1, (file_number + 1) * 500)
        if not os.path.exists(file_name):
            # 如果文件不存在，则创建
            generate_excel_file(file_name, [select_sheet_name])
            data_row = 1
        write_excel = load_workbook(file_name)
        write_sheet = write_excel[select_sheet_name]
        for y in range(1, sheet.max_column + 1):
            # 取数然后复制数据
            if int(sheet.cell(row=x, column=1).value) % 500 == 1:
                # 是每一个表格的第一行 需要添加表头
                value = sheet.cell(1, y).value
                cell = write_sheet.cell(1, y)
                cell.value = value
            value = sheet.cell(x, y).value
            cell = write_sheet.cell(data_row + 1, y)
            cell.value = value
        data_row += 1
        if int(sheet.cell(row=x, column=1).value) % 500 == 0:
            write_excel.save(file_name)

@time_counter()
def read_excel_file_to_generate_sub_file(file_path, fold_name, file_name, select_sheet_name, sub_file_row, sub_file_name_format):
    """
        @desc: 读取一个数据量很大的 excel 文件，将其拆分为多个 较小数据量的excel文件
        @thinking:  思路是使用pd来进行文件编辑操作，用load_workbook来进行文件读取操作
                    因为pd进行文件创建的话是比较快捷的一种方式，并不会耗费多少性能，因为数据需要提前汇总好格式
                    而load_workbook来进行文件写入操作或者创建操作的话，需要file.save()来进行保存，比较耗费性能
        @params: file_path:             大文件所在的路径
                 fold_name:             生成文件所在的文件夹
                 file_name:             大文件的文件名
                 select_sheet_name:     大文件中的sheet页
                 sub_file_row:          子文件允许的最大行
                 sub_file_name_format:  生成的子文件的命名规范(文件规范 % 明细行范围)
        @return: None                   无返回值
        @note: 38031.41403198242 ms
    """
    # 将大文件excel 拆分为多个 excel
    generate_dir = file_path + '\\' + fold_name
    if not os.path.exists(generate_dir):
        # 生成文件不存在，创建
        os.mkdir(generate_dir)
    file = load_workbook(file_path + '\\' + file_name)
    sheet = file[select_sheet_name]
    data_dict = {}
    for column in range(1, sheet.max_column):
        key = sheet.cell(1, column).value
        data_dict[key] = []
    for column in range(1, sheet.max_column):
        for row in range(2, sheet.max_row):
            key = sheet.cell(1, column).value
            value_list = data_dict[key]
            value_list.append(sheet.cell(row, column).value)
            data_dict[key] = value_list
    for number in range(sheet.max_row // sub_file_row + 1):
        # 生成的文件数量
        # 这里最大 - 2 的原因，因为最大是最后一行的下一行，所以需要减一，然后因为第一行是标题 也需要减去
        gen_file_name = sub_file_name_format % (number * sub_file_row + 1, (number + 1) * sub_file_row if (number +1) * sub_file_row <= sheet.max_row else sheet.max_row - 2)
        generate_dict = {}
        for key, value in data_dict.items():
            generate_dict[key] = value[number * sub_file_row: (number + 1) * sub_file_row]
        df = pd.DataFrame(generate_dict)
        df.to_excel(generate_dir + gen_file_name, index=False)

@time_counter(show_config='s')
def read_excel_to_generate_sub_file(file_path, fold_name, file_name, select_sheet_name, sub_file_row, sub_file_name_format):
    """
        @desc: 使用pandas读取excel文件并生成特定数据量的excel
        @note: @note: 21058.8641166687 ms
    """
    generate_dir = file_path + '\\' + fold_name
    if not os.path.exists(generate_dir):
        # 生成文件不存在，创建
        os.mkdir(generate_dir)
    origin_data = pd.read_excel(file_path + '\\' + file_name)
    title_list = list(origin_data)
    for number in range(len(origin_data) // sub_file_row + 1):
        generate_dict = dict()
        gen_file_name = sub_file_name_format % (number * sub_file_row + 1, (number + 1) * sub_file_row if (number +1) * sub_file_row <= len(origin_data) else len(origin_data))
        for key in title_list:
            generate_dict[key] = list(origin_data.get(key))[number * sub_file_row: (number + 1) * sub_file_row]
        df = pd.DataFrame(generate_dict)
        df.to_excel(generate_dir + gen_file_name, index=False)

def generate_excel_file(file_name, sheet_name_list):
    """
        @desc: 生成excel文件
    """
    workbook = xlsxwriter.Workbook(file_name)
    for sheet_name in sheet_name_list:
        # 遍历 sheet_name 创建sheet页
        work_sheet = workbook.add_worksheet(sheet_name)
    workbook.close()

read_excel_file_to_generate_sub_file(r'D:\back\Download', 'generate_file', 'all_summary.xlsx', '库存初始明细', 200, '\\模板%s-%s.xlsx')
read_excel_to_generate_sub_file(r'D:\back\Download', 'generate_file_pd', 'all_summary.xlsx', '库存初始明细', 200, '\\模板%s-%s.xlsx')

# read_excel_file(r'D:\back\Download', 'generate_file', 'all_summary.xlsx', '库存初始明细')
# import pandas as pd
# data = pd.read_excel(r"D:\back\Download\all_summary.xlsx", )
# # pd.read_excel(excel_name, sheet_name)
# # 指定读取特定文件的特定sheet页
# # 读取 导入序号 列的值在 1000 - 2000 的数据
# # data.loc[(data['*导入序号'] > 1000) & (data['*导入序号'] < 2000), list(data)]
# print(data)